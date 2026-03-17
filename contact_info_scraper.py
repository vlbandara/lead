import logging
import argparse
from concurrent.futures import ThreadPoolExecutor
import requests
import re
import json
import time
import random
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

import os
from bs4 import BeautifulSoup
from urllib.parse import urlparse, urljoin, urlencode, parse_qs
from googlesearch import search
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ── OpenRouter AI config ──────────────────────────────────────────────────────
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "")
OPENROUTER_MODEL   = "minimax/minimax-m2.5"
OPENROUTER_URL     = "https://openrouter.ai/api/v1/chat/completions"

# ── Browser fallback (for sites that block requests with 403) ──────────────────
USE_BROWSER_FALLBACK = os.getenv("USE_BROWSER_FALLBACK", "").strip().lower() in ("1", "true", "yes")
_BROWSER_BLOCKING_CODES = (403, 401, 405, 429)  # try headless browser when these occur

# ── Proxy config ──────────────────────────────────────────────────────────────
def load_proxies(proxy_file: str = "proxies.txt") -> list:
    """Load proxies from file (one per line, format: http://user:pass@host:port or just http://host:port)"""
    proxies = []
    if os.path.exists(proxy_file):
        with open(proxy_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    proxies.append(line)
    return proxies

PROXY_LIST = load_proxies()
PROXY_INDEX = 0

def get_next_proxy():
    """Return next proxy as requests format dict, or None if no proxies available"""
    global PROXY_INDEX
    if not PROXY_LIST:
        return None
    proxy = PROXY_LIST[PROXY_INDEX % len(PROXY_LIST)]
    PROXY_INDEX += 1
    return {"http": proxy, "https": proxy}

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("log.txt"),
        logging.StreamHandler()
    ]
)

# Tracking/analytics query parameters to strip from stored URLs
_TRACKING_PARAMS = {
    'utm_source', 'utm_medium', 'utm_campaign', 'utm_term', 'utm_content',
    'srsltid', 'fbclid', 'gclid', 'gad_source', 'msclkid', 'twclid',
    'igshid', 'ref', 'source', '_ga', 'mc_eid',
}

def clean_url(url: str) -> str:
    """Strip tracking query parameters and return a clean URL."""
    parsed = urlparse(url)
    clean_qs = {k: v for k, v in parse_qs(parsed.query).items()
                if k.lower() not in _TRACKING_PARAMS}
    cleaned = parsed._replace(query=urlencode(clean_qs, doseq=True))
    return cleaned.geturl().rstrip('?')

def get_timestamp():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def remove_duplicates(data):
    seen = set()
    result = []
    for item in data:
        item = item.strip()
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result

# ── Regex helpers (kept as first-pass / fallback) ─────────────────────────────

def decode_cloudflare_email(encoded_hex: str) -> str:
    """Decode a Cloudflare-obfuscated email from its data-cfemail hex string."""
    try:
        r = int(encoded_hex[:2], 16)
        return ''.join(
            chr(int(encoded_hex[i:i+2], 16) ^ r)
            for i in range(2, len(encoded_hex), 2)
        )
    except Exception:
        return ""

def extract_cf_protected_emails(soup) -> set:
    """Find all Cloudflare-protected emails in a BeautifulSoup tree."""
    emails = set()
    for tag in soup.select('a[data-cfemail]'):
        decoded = decode_cloudflare_email(tag['data-cfemail'])
        if decoded and '@' in decoded:
            emails.add(decoded)
    return emails

def extract_mailto_emails(soup) -> set:
    """Extract emails from <a href="mailto:..."> links — missed by plain text regex."""
    emails = set()
    for tag in soup.select('a[href^="mailto:"]'):
        raw = tag['href'][len('mailto:'):].split('?')[0].strip()
        if raw and '@' in raw:
            emails.add(raw)
    return emails

def extract_jsonld_contacts(soup) -> tuple[set, set]:
    """
    Extract emails and phones from JSON-LD structured data blocks.
    Modern sites embed reliable contact info here for search engines.
    """
    emails, phones = set(), set()
    email_re = re.compile(r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}')

    def _walk(obj):
        if isinstance(obj, dict):
            for key, val in obj.items():
                if key.lower() in ('email', 'e-mail') and isinstance(val, str) and '@' in val:
                    emails.add(val.strip())
                elif key.lower() in ('telephone', 'phone', 'faxnumber') and isinstance(val, str):
                    phones.add(val.strip())
                else:
                    _walk(val)
        elif isinstance(obj, list):
            for item in obj:
                _walk(item)

    for script in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(script.string or '')
            _walk(data)
        except Exception:
            # Fallback: regex scan the raw script text
            raw = script.string or ''
            emails.update(email_re.findall(raw))

    return emails, phones

def get_email(html):
    pattern = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    return set(re.findall(pattern, html))

def _phone_has_min_digits(phone: str, min_digits: int = 7) -> bool:
    """Return True if phone string has at least min_digits numeric characters."""
    return len(re.sub(r'\D', '', phone)) >= min_digits


def get_phone(html):
    phone_patterns = [
        r"(\+?\d{1,3}[-.\s]?)?\(?\d{2,3}?\)?[-.\s]?\d{3}[-.\s]?\d{4}",
        r"\d{2,3}[-.\s]?\d{3}[-.\s]?\d{4}",
        r"((?:\d{2,3}|\(\d{2,3}\))?(?:\s|-|\.)?\d{3,4}(?:\s|-|\.)?\d{4})",
        r"\d{4}[-.\s]?\d{3}[-.\s]?\d{3}",
        r"(\+\d{1,3}[- ]?)?\(?\d{1,4}?\)?[- ]?\d{1,4}[- ]?\d{1,4}[- ]?\d{1,4}",
    ]
    raw = set(num for pattern in phone_patterns for num in re.findall(pattern, html))
    return set(p for p in raw if _phone_has_min_digits(p))


def normalize_phone(phone: str) -> str:
    """Normalize phone number to a standard format."""
    digits = re.sub(r'\D', '', phone)
    
    # Sri Lankan mobile: 07X XXX XXXX (10 digits starting with 7)
    if len(digits) == 10 and digits.startswith('7'):
        return f"+94 {digits[0:2]} {digits[2:5]} {digits[5:]}"
    
    # International: 9477 XXXX XXXX (11 digits starting with 947)
    if len(digits) == 11 and digits.startswith('947'):
        return f"+94 {digits[3:5]} {digits[5:8]} {digits[8:]}"
    
    # Already has +94 prefix with space: +94 77 XXX XXXX
    if phone.startswith('+94'):
        return phone
    
    # US/other: keep as-is if short
    if len(digits) <= 10:
        return phone
    return phone


def dedup_and_normalize_phones(phones: set) -> list:
    """Deduplicate and normalize phone numbers."""
    # First normalize all phones
    normalized = []
    for phone in phones:
        norm = normalize_phone(phone)
        normalized.append(norm)
    
    # Now deduplicate by comparing digit-only versions
    seen_digits = set()
    unique_phones = []
    for phone in normalized:
        digits = re.sub(r'\D', '', phone)
        if digits and digits not in seen_digits:
            seen_digits.add(digits)
            unique_phones.append(phone)
    
    return sorted(unique_phones)


# ── OpenRouter LLM helpers ────────────────────────────────────────────────────

def _call_llm(prompt: str, url: str = "", max_tokens: int = 600) -> str:
    """Send a prompt to OpenRouter and return the raw response text."""
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
        "HTTP-Referer": url or "https://github.com/ContactInfoScraper",
        "X-Title": "ContactInfoScraper",
    }
    payload = {
        "model": OPENROUTER_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
        "max_tokens": max_tokens,
    }
    try:
        proxies = get_next_proxy()
        res = requests.post(OPENROUTER_URL, headers=headers, json=payload, timeout=30, proxies=proxies)
        res.raise_for_status()
        content = res.json()["choices"][0]["message"].get("content")
        return (content or "").strip()
    except Exception as e:
        logging.warning(f"{get_timestamp()} - LLM call failed: {e}")
        return ""


def llm_extract_contacts(page_text: str, url: str) -> tuple[list, list, str]:
    """
    Ask the LLM to extract emails, phones, and a brief business summary
    from raw page text. Returns (emails, phones, summary_note).
    """
    truncated = page_text[:5000] if len(page_text) > 5000 else page_text

    prompt = f"""You are a precise contact information extraction assistant.

Extract ALL email addresses and phone numbers from the webpage text below (from: {url}).

Return ONLY valid JSON — no markdown, no explanation:
{{
  "emails": ["email@domain.com"],
  "phones": ["+1-800-555-0100"],
  "summary": "One sentence describing the business and contact type found"
}}

Rules:
- Include ONLY real, plausible emails (skip example.com, placeholder, w3.org, schemata, etc.)
- IGNORE any text like "[email protected]" or "[email protected]" — these are Cloudflare obfuscation placeholders, not real emails
- Include ONLY real phone numbers — NOT version numbers, zip codes, IDs, or dates
- Copy phone numbers EXACTLY as they appear in the text — do NOT reformat, do NOT add/guess country codes
- Return empty arrays if nothing valid is found
- Keep the summary to one concise sentence

Webpage text:
{truncated}"""

    raw = _call_llm(prompt, url=url, max_tokens=600)
    if not raw:
        return [], [], ""

    try:
        match = re.search(r'\{.*\}', raw, re.DOTALL)
        if match:
            data = json.loads(match.group())
            return (
                [e.strip() for e in data.get("emails", []) if e.strip()],
                [p.strip() for p in data.get("phones", []) if p.strip()],
                data.get("summary", ""),
            )
    except Exception as e:
        logging.warning(f"{get_timestamp()} - LLM JSON parse error: {e} | raw: {raw[:200]}")

    return [], [], ""


def llm_validate_and_merge(
    regex_emails: list, regex_phones: list,
    ai_emails: list, ai_phones: list,
    url: str,
    jsonld_emails: set = None,
    jsonld_phones: set = None,
) -> tuple[list, list]:
    """
    Ask the LLM to merge regex + AI results and remove false positives.
    Returns cleaned (emails, phones).
    """
    # Track JSON-LD sources for fallback
    if jsonld_emails is None:
        jsonld_emails = set()
    if jsonld_phones is None:
        jsonld_phones = set()
    
    all_emails = list(set(regex_emails) | set(ai_emails))
    all_phones = list(set(str(p) for p in regex_phones) | set(ai_phones))

    if not all_emails and not all_phones:
        return [], []

    prompt = f"""You are a contact data quality expert. Your ONLY job is to decide which candidates are real — you must NOT reformat or modify any string.

Candidate emails: {all_emails}
Candidate phones: {all_phones}

Return ONLY valid JSON — no markdown, no explanation:
{{
  "emails": ["copy exact strings from candidate emails that are real"],
  "phones": ["copy exact strings from candidate phones that are real"]
}}

Strict rules — READ CAREFULLY:
- Copy each kept value EXACTLY as it appears in the candidates list — zero modifications, zero reformatting
- REMOVE emails that are clearly fake (schema URLs, example.com, escaped HTML, file paths)
- KEEP any plausible business email: info@, contact@, support@, sales@, noreply@, etc.
- REMOVE phone numbers that are version strings (e.g. 1.0.0), years (2024), zip codes, or very short IDs (< 6 digits)
- KEEP all other phone numbers, copied verbatim — do NOT add country codes, do NOT change formatting
- Deduplicate by keeping the first occurrence"""

    raw = _call_llm(prompt, url=url, max_tokens=400)
    if not raw:
        return all_emails, all_phones

    try:
        match = re.search(r'\{.*\}', raw, re.DOTALL)
        if match:
            data = json.loads(match.group())
            return (
                [e.strip() for e in data.get("emails", all_emails) if e.strip()],
                [p.strip() for p in data.get("phones", all_phones) if p.strip()],
            )
    except Exception as e:
        logging.warning(f"{get_timestamp()} - LLM validation parse error: {e}")

    # Fallback: if LLM returned empty/bad but we had JSON-LD data, preserve it
    final_emails = [e.strip() for e in data.get("emails", []) if e.strip()] if match else []
    final_phones = [p.strip() for p in data.get("phones", []) if p.strip()] if match else []
    
    if (not final_emails and jsonld_emails) or (not final_phones and jsonld_phones):
        logging.info(f"{get_timestamp()} - LLM returned empty, preserving JSON-LD data")
        final_emails = list(jsonld_emails) if not final_emails else final_emails
        final_phones = list(jsonld_phones) if not final_phones else final_phones
    
    return final_emails, final_phones

# ── Core scraping helpers ─────────────────────────────────────────────────────

_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_3_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0",
]

def _browser_headers(url: str) -> dict:
    """Return realistic browser-like headers for the given URL."""
    parsed = urlparse(url)
    origin = f"{parsed.scheme}://{parsed.netloc}"
    return {
        "User-Agent": random.choice(_USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate",
        "Referer": origin,
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Cache-Control": "max-age=0",
    }


def _fetch_with_browser(url: str):
    """
    Fetch a URL using a headless browser (Playwright). Use when the site blocks
    normal HTTP requests (e.g. 403 Forbidden). Returns a response-like object
    with .text and .url, or None on failure.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logging.warning(f"{get_timestamp()} - Playwright not installed; run: pip install playwright && playwright install chromium")
        return None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent=random.choice(_USER_AGENTS),
                viewport={"width": 1280, "height": 720},
                ignore_https_errors=True,
            )
            page = context.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=25000)
            final_url = page.url
            html = page.content()
            context.close()
            browser.close()
        # Response-like object so downstream code can use .text and .url
        class _BrowserResponse:
            pass
        resp = _BrowserResponse()
        resp.text = html
        resp.url = final_url
        logging.info(f"{get_timestamp()} - Fetched via browser fallback: {url}")
        return resp
    except Exception as e:
        logging.warning(f"{get_timestamp()} - Browser fetch failed for {url}: {e}")
        return None


def _is_blocking_http_error(exc: requests.exceptions.RequestException) -> bool:
    """True if the error is an HTTP response that might be bypassed by a real browser (e.g. 403)."""
    if not isinstance(exc, requests.exceptions.HTTPError):
        return False
    if getattr(exc, "response", None) is None:
        return False
    return getattr(exc.response, "status_code", None) in _BROWSER_BLOCKING_CODES


def log_no_results(info_type, source):
    msg = f'{get_timestamp()} - No {info_type} found on {source}.'
    logging.info(msg)
    print(msg)

def fetch_data_with_error_handling(url, headers=None, max_retries: int = 3):
    session = requests.Session()
    session.headers.update({"Accept-Encoding": "gzip, deflate"})
    last_exception = None
    for attempt in range(max_retries):
        try:
            req_headers = _browser_headers(url)
            if headers:
                req_headers.update(headers)
            proxies = get_next_proxy()
            res = session.get(url, headers=req_headers, timeout=20, allow_redirects=True, proxies=proxies)
            res.raise_for_status()
            if hasattr(res, 'content'):
                res.text  # trigger decompression
            return res
        except requests.exceptions.RequestException as e:
            last_exception = e
            delay = 2 ** attempt
            logging.warning(f"{get_timestamp()} - Error accessing URL: {url}, Error: {e} (attempt {attempt + 1}/{max_retries}, retry in {delay}s)")
            if attempt < max_retries - 1:
                time.sleep(delay)
    # All retries failed: try browser fallback when site blocks with 403 etc.
    if USE_BROWSER_FALLBACK and last_exception and _is_blocking_http_error(last_exception):
        logging.info(f"{get_timestamp()} - Trying browser fallback for blocked URL: {url}")
        return _fetch_with_browser(url)
    return None

# Patterns that indicate a dedicated contact page by URL path
_CONTACT_HREF_RE = re.compile(
    r'/(contact[-_]?us?|pages/contact|reach[-_]?us?|get[-_]in[-_]touch'
    r'|reach[-_]out|write[-_]to[-_]us|about/contact|help/contact'
    r'|support/contact|enquir|inquiry|enquiry)',
    re.IGNORECASE,
)
# Patterns that indicate a contact link by visible text
_CONTACT_TEXT_RE = re.compile(
    r'\b(contact|reach us|get in touch|reach out|write to us|email us|call us|enquir|inquiry)\b',
    re.IGNORECASE,
)

def find_contact_pages(soup, base_url: str) -> list:
    """
    Discover dedicated contact page URLs by matching either the href path
    or the visible link text. Returns deduplicated absolute URLs.
    """
    seen = set()
    results = []
    for a in soup.find_all('a', href=True):
        href = a['href'].strip()
        if not href or href.startswith('mailto:') or href.startswith('#'):
            continue
        text = a.get_text(separator=' ', strip=True)
        if _CONTACT_HREF_RE.search(href) or _CONTACT_TEXT_RE.search(text):
            abs_url = urljoin(base_url, href)
            # Drop fragment and trailing slashes for dedup
            abs_url = abs_url.split('#')[0].rstrip('/')
            if abs_url and abs_url not in seen and abs_url != base_url.rstrip('/'):
                seen.add(abs_url)
                results.append(abs_url)
    return results

def extract_facebook_url(soup):
    fb_links = [
        a['href'] for a in soup.find_all('a', href=re.compile('facebook.com', re.IGNORECASE))
        if 'href' in a.attrs
    ]
    return fb_links[0] if fb_links else None

def search_google(company_name):
    query = f"{company_name} phone number"
    try:
        return list(search(query, num_results=5))
    except Exception as e:
        logging.warning(f"{get_timestamp()} - Google search failed: {e}")
        return []

def get_phone_from_social_media(url):
    res = fetch_data_with_error_handling(url)
    if res:
        return get_phone(BeautifulSoup(res.text, 'lxml').get_text())
    return []

# ── Main gather function ──────────────────────────────────────────────────────

def gather_contact_info(url):
    res = fetch_data_with_error_handling(url)

    if not res:
        return None

    canonical_url = clean_url(res.url)

    print(f'\n{"="*50}')
    print(f'  Scraping: {canonical_url}')
    print(f'{"="*50}')
    logging.info(f'{get_timestamp()} - Scraping: {canonical_url}')

    soup = BeautifulSoup(res.text, 'lxml')
    landing_text = soup.get_text(separator=' ', strip=True)
    company_name = soup.title.string.strip() if soup.title else "Unknown"

    # ── Step 1: Find dedicated contact pages ──────────────────────────────────
    print("  [1/3] Running regex extraction...")
    contact_pages = find_contact_pages(soup, canonical_url)

    # primary_text is what gets sent to the AI — prefer the contact page
    primary_text = landing_text
    primary_url  = canonical_url

    ld_emails, ld_phones = extract_jsonld_contacts(soup)
    if ld_emails or ld_phones:
        logging.info(f'{get_timestamp()} - JSON-LD contacts on landing page: emails={ld_emails}, phones={ld_phones}')

    regex_emails = (list(get_email(landing_text))
                    + list(extract_cf_protected_emails(soup))
                    + list(extract_mailto_emails(soup))
                    + list(ld_emails))
    regex_phones = list(get_phone(landing_text)) + list(ld_phones)

    if contact_pages:
        logging.info(f'{get_timestamp()} - Found contact page(s): {contact_pages}')
    else:
        logging.info(f'{get_timestamp()} - No dedicated contact page found; using landing page.')

    for cp_url in contact_pages:
        sub = fetch_data_with_error_handling(cp_url)
        if not sub:
            continue
        sub_soup = BeautifulSoup(sub.text, 'lxml')
        sub_text = sub_soup.get_text(separator=' ', strip=True)

        print(f"     → Contact page: {cp_url}")
        logging.info(f'{get_timestamp()} - Scraped contact page: {cp_url}')

        sub_ld_emails, sub_ld_phones = extract_jsonld_contacts(sub_soup)
        if sub_ld_emails or sub_ld_phones:
            logging.info(f'{get_timestamp()} - JSON-LD contacts on {cp_url}: emails={sub_ld_emails}, phones={sub_ld_phones}')
        regex_emails += list(get_email(sub_text))
        regex_emails += list(extract_cf_protected_emails(sub_soup))
        regex_emails += list(extract_mailto_emails(sub_soup))
        regex_emails += list(sub_ld_emails)
        regex_phones += list(get_phone(sub_text)) + list(sub_ld_phones)

        # Use the first successfully fetched contact page as the AI primary source
        if primary_url == canonical_url:
            primary_text = sub_text
            primary_url  = clean_url(sub.url) if sub.url else cp_url

    # ── Step 2: AI extraction (on the best available page) ───────────────────
    print("  [2/3] Running AI extraction...")
    ai_emails, ai_phones, ai_summary = llm_extract_contacts(primary_text, primary_url)
    logging.info(f"{get_timestamp()} - AI extracted from {primary_url}: emails={ai_emails}, phones={ai_phones}")

    # ── Step 3: AI validation & merge ─────────────────────────────────────────
    print("  [3/3] AI validating & merging results...")
    final_emails, final_phones = llm_validate_and_merge(
        regex_emails, regex_phones, ai_emails, ai_phones, res.url,
        jsonld_emails=ld_emails, jsonld_phones=ld_phones
    )

    # ── Fallback chain if phones still empty ─────────────────────────────────
    if not final_phones:
        facebook_url = extract_facebook_url(soup)
        if facebook_url:
            fb_phones = list(get_phone_from_social_media(facebook_url))
            if fb_phones:
                final_phones.extend(fb_phones)
            else:
                log_no_results('phone numbers', 'Facebook')

    if not final_phones:
        maps_results = search_google(f"{company_name} site:maps.google.com")
        for maps_url in maps_results:
            maps_phones = list(get_phone_from_social_media(maps_url))
            if maps_phones:
                final_phones.extend(maps_phones)
                break
        if not final_phones:
            log_no_results('phone numbers', 'Google Maps')

    if not final_phones:
        yelp_results = search_google(f"{company_name} site:yelp.com")
        for yelp_url in yelp_results:
            yelp_phones = list(get_phone_from_social_media(yelp_url))
            if yelp_phones:
                final_phones.extend(yelp_phones)
                break
        if not final_phones:
            log_no_results('phone numbers', 'Yelp')

    if not final_phones:
        for result in search_google(company_name):
            g_res = fetch_data_with_error_handling(result)
            if g_res:
                g_phones = list(get_phone(g_res.text))
                g_emails = list(get_email(g_res.text))
                if g_phones:
                    final_phones.extend(g_phones)
                    final_emails.extend(g_emails)
                    break
            else:
                log_no_results('phone numbers', 'Google')

    final_emails = remove_duplicates(final_emails)
    final_phones = [p for p in remove_duplicates(final_phones) if _phone_has_min_digits(p)]

    return {
        'Website': canonical_url,
        'Email': final_emails,
        'Phone': final_phones,
        'AI Summary': ai_summary,
    }

# ── Excel output ──────────────────────────────────────────────────────────────

def save_to_excel(contacts, output_path: str = "contacts.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = ['Website', 'Email', 'Phone', 'AI Summary']
    header_fill   = PatternFill("solid", fgColor="2E4057")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    alt_fill      = PatternFill("solid", fgColor="F0F4F8")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, contact in enumerate(contacts, 2):
        ws.cell(row=row_idx, column=1, value=contact['Website'])
        ws.cell(row=row_idx, column=2, value=', '.join(contact['Email']) or 'None')
        ws.cell(row=row_idx, column=3, value=', '.join(contact['Phone']) or 'None')
        ws.cell(row=row_idx, column=4, value=contact.get('AI Summary', ''))
        if row_idx % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = alt_fill
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    col_widths = [40, 45, 35, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    logging.info(f'{get_timestamp()} - Saved results to {output_path}')


def save_to_json(contacts, output_path: str = "contacts.json"):
    """Save contacts to JSON file."""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(contacts, f, indent=2, ensure_ascii=False)
    logging.info(f'{get_timestamp()} - Saved results to {output_path}')


# ── Entry point ───────────────────────────────────────────────────────────────

def _normalize_url(url: str) -> str:
    url = url.strip()
    if not url:
        return ""
    if not urlparse(url).scheme:
        url = "https://" + url
    return url


def _is_valid_url(url: str) -> bool:
    return bool(re.match(r"https?://[^\s/$.?#].[^\s]*", url))


def _load_urls_from_file(path: str) -> list:
    """Load and validate URLs from a file; return list of valid URLs."""
    urls = []
    p = Path(path)
    if not p.exists():
        return urls
    for line in p.read_text(encoding="utf-8", errors="ignore").splitlines():
        u = _normalize_url(line)
        if u and _is_valid_url(u):
            urls.append(u)
    return urls


def main():
    parser = argparse.ArgumentParser(description="ContactInfoScraper — extract emails and phones from websites.")
    parser.add_argument(
        "-f", "--file",
        default="web_urls.txt",
        help="Read URLs from this file (default: web_urls.txt); used if file exists",
    )
    parser.add_argument(
        "-o", "--output",
        default="contacts.xlsx",
        help="Output Excel path (default: contacts.xlsx)",
    )
    parser.add_argument(
        "-i", "--interactive",
        action="store_true",
        help="Prompt for URLs interactively; new URLs are appended to the file",
    )
    parser.add_argument(
        "-w", "--workers",
        type=int,
        default=3,
        help="Number of concurrent scrape workers (default: 3)",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Also save output as JSON file",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="Request timeout in seconds (default: 30)",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose logging",
    )
    parser.add_argument(
        "--proxy-file",
        default="proxies.txt",
        help="File containing proxy list (one per line)",
    )
    args = parser.parse_args()

    # Load proxies if file exists
    global PROXY_LIST
    if args.proxy_file and os.path.exists(args.proxy_file):
        PROXY_LIST = load_proxies(args.proxy_file)
        logging.info(f"{get_timestamp()} - Loaded {len(PROXY_LIST)} proxies from {args.proxy_file}")

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    print("\n" + "=" * 50)
    print("  ContactInfoScraper  —  AI-Enhanced Edition")
    print("=" * 50 + "\n")

    urls = _load_urls_from_file(args.file)
    if urls:
        logging.info(f"{get_timestamp()} - Loaded {len(urls)} URL(s) from {args.file}")

    if args.interactive:
        while True:
            url = input('Enter a URL (or "n" to start scraping): ').strip()
            if url.lower() == "n":
                break
            u = _normalize_url(url)
            if not _is_valid_url(u):
                print("  Invalid URL — please try again.\n")
                logging.warning(f"{get_timestamp()} - Invalid URL: {url}")
                continue
            urls.append(u)
            with open(args.file, "a", encoding="utf-8") as f:
                f.write(u + "\n")
            print(f"  Added: {u}\n")

    if not urls:
        print("No URLs to scrape. Use -i for interactive input or add URLs to your --file.")
        logging.info(f"{get_timestamp()} - No URLs to scrape.")
        return

    if args.workers <= 1:
        contacts = []
        for url in urls:
            contact_info = gather_contact_info(url)
            if contact_info:
                contacts.append(contact_info)
                logging.info(f"{get_timestamp()} - Result: {json.dumps(contact_info, indent=2)}")
                print(f"\n  {'─'*46}")
                print(f"  Website : {contact_info['Website']}")
                print(f"  Emails  : {', '.join(contact_info['Email']) or 'None'}")
                print(f"  Phones  : {', '.join(contact_info['Phone']) or 'None'}")
                print(f"  AI Note : {contact_info['AI Summary'] or '—'}")
                print(f"  {'─'*46}\n")
            time.sleep(1)
    else:
        contacts = []
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            for contact_info in executor.map(gather_contact_info, urls):
                if contact_info:
                    contacts.append(contact_info)
                    logging.info(f"{get_timestamp()} - Result: {json.dumps(contact_info, indent=2)}")
                    print(f"\n  {'─'*46}")
                    print(f"  Website : {contact_info['Website']}")
                    print(f"  Emails  : {', '.join(contact_info['Email']) or 'None'}")
                    print(f"  Phones  : {', '.join(contact_info['Phone']) or 'None'}")
                    print(f"  AI Note : {contact_info['AI Summary'] or '—'}")
                    print(f"  {'─'*46}\n")

    # Apply phone deduplication to all contacts
    for contact in contacts:
        contact['Phone'] = dedup_and_normalize_phones(set(contact.get('Phone', [])))

    save_to_excel(contacts, args.output)
    if args.json:
        json_path = args.output.replace('.xlsx', '.json')
        save_to_json(contacts, json_path)
    print(f"\n  Done! Results saved to {args.output}\n")


if __name__ == "__main__":
    main()
